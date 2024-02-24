import discord
from discord.ext import commands
from random import randint
import re
import openpyxl
from tenor import get_gif_pokemon
from weather import get_weather
from pytube import YouTube
import asyncio
import utils.utils as utils
from datetime import datetime

bot = commands.Bot(command_prefix="!",
                   intents=discord.Intents.all(),
                   status=discord.Status.idle)

# Sync bot commands with discord

@bot.event
async def on_ready():
    print(f"We have logged in as {bot.user}")

    try:
        synced = await bot.tree.sync()
        print(f"Synced {len(synced)} command(s)")
    except Exception as e:
        print(e)

# "oi" command to make bot say hi back to who used the command

@bot.tree.command(name="oi", description="Bot te diz oi de volta")
async def hello(interaction: discord.Interaction):
    # ephemeral=True)
    await interaction.response.send_message(f"Oi, {interaction.user.mention}!")

# Make bot say something specific

@bot.tree.command(name="diga", description="Bot diz algo específico")
async def hello(interaction: discord.Interaction, o_que_falar: str):
    await interaction.response.send_message(f"{o_que_falar}!")

# Roll dice command using regex

@bot.tree.command(name="rd", description="Rola um dado de 20 lados")
async def rd(interaction: discord.Interaction, dice: str):
    pattern = re.compile("\\b([1-9][0-9]?|100)d([1-9][0-9]?|100)\\b")
    print(dice)
    print(pattern.search(dice))
    if (pattern.search(dice)):
        values_list = dice.split("d")
        dice_quantity = values_list[0]
        dice_type = values_list[1]
        list_dices = []

        for i in range(int(dice_quantity)):
            list_dices.append(randint(1, int(dice_type)))
        str_list_dices = [str(i) for i in list_dices]
        result_dices = ", ".join(str_list_dices)
        total_value_dices = sum(list_dices)

        await interaction.response.send_message(f"Você rolou {dice_quantity}d{dice_type}: {result_dices}\nTotal:({total_value_dices})")
    else:
        await interaction.response.send_message(f"Você digitou algo errado, tente algo como \"2d20\" (2 dados de 20 lados) sem aspas\n"
                                                f"Lembre-se de utilizar apenas valores entre 1 e 100", ephemeral=True)

# Get a random pokemon command

workbook_pokemons = openpyxl.load_workbook("src/assets/pokemons.xlsx")["Planilha1"]

@bot.tree.command(name="pokemon", description="Escolhe um pokemon aleatório para você")
async def pokemon(interaction: discord.Interaction):
    number_of_rows = 1
    for _ in workbook_pokemons.iter_rows(min_row=2):
        number_of_rows += 1
    
    pokemon_choice = randint(2, number_of_rows)
    pokemon_number = workbook_pokemons[pokemon_choice][0].value
    pokemon_gen = workbook_pokemons[pokemon_choice][1].value
    pokemon_name = workbook_pokemons[pokemon_choice][2].value
    pokemon_type = workbook_pokemons[pokemon_choice][3].value
    pokemon_info = workbook_pokemons[pokemon_choice][4].value
    pokemon_sprite = workbook_pokemons[pokemon_choice][5].value
    pokemon_gif = get_gif_pokemon(pokemon_name)

    embed = discord.Embed(title="Pokemon")
    embed.set_author(name=f"Pokemon do {interaction.user.name}", icon_url=interaction.user.avatar)
    embed.set_image(url=pokemon_gif)
    embed.set_thumbnail(url=f"https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/{pokemon_sprite}.png")
    embed.add_field(name="**Nome**:", value=pokemon_name)
    embed.add_field(name="**Nº**:", value=pokemon_number)
    embed.add_field(name="**Geração**:", value=f"{pokemon_gen}ª")
    embed.add_field(name="**Tipo**:", value=pokemon_type)
    embed.add_field(name="**Descrição**:", value=pokemon_info, inline=False)
    if pokemon_number == "0376":
        embed.add_field(name="**Vale a pena trocar metagross male por?**", value="Ah, já troquei")

    await interaction.response.send_message(embed=embed)

# Get a specific pokemon command

@bot.tree.command(name="achar_pokemon", description="Pesquise seu pokemon")
async def find_pokemon(interaction: discord.Interaction, pokemon: str):
    try:
        for row in workbook_pokemons.iter_rows(min_row=2):
            if pokemon.lower() == row[2].value.lower():
                pokemon_number = row[0].value
                pokemon_gen = row[1].value
                pokemon_name = row[2].value
                pokemon_type = row[3].value
                pokemon_info = row[4].value
                pokemon_sprite = row[5].value
                pokemon_gif = get_gif_pokemon(pokemon_name)
                break
        
        embed = discord.Embed(title="Pokemon")
        embed.set_author(name=f"Pokemon do {interaction.user.name}", icon_url=interaction.user.avatar)
        embed.set_image(url=pokemon_gif)
        embed.set_thumbnail(url=f"https://raw.githubusercontent.com/PokeAPI/sprites/master/sprites/pokemon/{pokemon_sprite}.png")
        embed.add_field(name="**Nome**:", value=pokemon_name)
        embed.add_field(name="**Nº**:", value=pokemon_number)
        embed.add_field(name="**Geração**:", value=f"{pokemon_gen}ª")
        embed.add_field(name="**Tipo**:", value=pokemon_type)
        embed.add_field(name="**Descrição**:", value=pokemon_info, inline=False)
        if pokemon_number == "0376":
            embed.add_field(name="**Vale a pena trocar metagross male por?**", value="Ah, já troquei")
        
        await interaction.response.send_message(embed=embed)
    except Exception as e:
        print(e)
        await interaction.response.send_message("Não achei seu pokemon, tente novamente", delete_after=15)


# Weather bot command    

@bot.tree.command(name="tempo", description="Bot mostra previsão do tempo do local solicitado")
async def weather(interaction: discord.Interaction, city: str):
    try:
        weather = get_weather(city)
        location = weather["name"]
        description = weather["weather"][0]["description"].capitalize()
        minimum_temperature = weather["main"]["temp_min"]
        maximum_temperature = weather["main"]["temp_max"]
        feels_like = weather["main"]["feels_like"]
        icon = weather["weather"][0]["icon"]

        embed = discord.Embed(title=f"Previsão do tempo para {location}")
        embed.set_author(name=interaction.user.name, icon_url=interaction.user.avatar)
        embed.set_thumbnail(url=f"https://openweathermap.org/img/wn/{icon}@2x.png")
        embed.add_field(name="**Temperatura mínima:**", value=f"{minimum_temperature}ºC")
        embed.add_field(name="**Temperatura máxima:**", value=f"{maximum_temperature}ºC")
        embed.add_field(name="**Sensação térmica:**", value=f"{feels_like}ºC", inline=False)
        embed.add_field(name="**Descrição:**", value=description, inline=False)

        await interaction.response.send_message(embed=embed)
    except Exception as e:
        print(e)
        await interaction.response.send_message("Não achei esta cidade, tente novamente", delete_after=15)

# Music command

music_queue = []
musics_titles = []

@bot.tree.command(name="tocar", description="Bot toca o áudio de uma url do youtube")
async def play(interaction: discord.Interaction, titulo_ou_url: str):
    if not discord.utils.get(bot.voice_clients, guild=interaction.guild):
        channel = interaction.user.voice.channel
        voice_client = await channel.connect(self_deaf=True)
    else:
        voice_client = discord.utils.get(bot.voice_clients)
    
    await interaction.response.defer()
    
    url = utils.get_video_url_by_title(titulo_ou_url)

    # Download youtube mp3 audio
    yt = YouTube(url)
    
    if yt.length > 900:
        msg = await interaction.followup.send("Música muito longa")
        return 0
    
    msg = await interaction.followup.send(f"Música {yt.title} adicionada a fila")
    
    asyncio.create_task(utils.delete_message(msg))

    musics_titles.append(yt.title)

    clean_title = datetime.now().microsecond

    music_queue.append(clean_title)
    print(f"{clean_title} added to music_queue")
    print(f"current music queue: {music_queue}")

    audio_stream = yt.streams.filter(only_audio=True).first()
    audio_stream.download(filename=f"src/assets/audios/{music_queue[-1]}.mp3")

    # Start playing the queue
    if not voice_client.is_playing():
        voice_client.play(discord.FFmpegPCMAudio(f"src/assets/audios/{music_queue[0]}.mp3"))
        
        await bot.change_presence(activity=discord.Activity(type=discord.ActivityType.listening, name=f"{musics_titles[0]}"))
        
        while music_queue:
            await asyncio.sleep(3)
            if not voice_client.is_playing() and not voice_client.is_paused() and voice_client.is_connected():
                utils.delete_music(music_queue[0])
                music_queue.pop(0)
                musics_titles.pop(0)
                print(f"current music queue: {music_queue}")
                if not music_queue:
                    break
                voice_client.play(discord.FFmpegPCMAudio(f"src/assets/audios/{music_queue[0]}.mp3"))
                await bot.change_presence(activity=discord.Activity(type=discord.ActivityType.listening, name=f"{musics_titles[0]}"))
        
        await bot.change_presence(status=discord.Status.idle, activity=None)
        await voice_client.disconnect()

@bot.tree.command(name="parar", description="Para de tocar")
async def stop(interaction: discord.Interaction):
    voice_client = discord.utils.get(bot.voice_clients, guild=interaction.guild)
    if voice_client and voice_client.is_playing():
        await voice_client.disconnect()
        utils.delete_queue(music_queue)
        music_queue.clear()
        await bot.change_presence(status=discord.Status.idle, activity=None)
        await interaction.response.send_message("Parando de tocar", delete_after=3)
    else:
        await interaction.response.send_message("Não há música tocando para parar", delete_after=3)

@bot.tree.command(name="pular", description="Pula a música atual")
async def skip(interaction: discord.Interaction):
    voice_client = discord.utils.get(bot.voice_clients, guild=interaction.guild)
    if voice_client and voice_client.is_playing():
        voice_client.stop()
        await interaction.response.send_message("Pulando música atual", delete_after=3)
    else:
        await interaction.response.send_message("Não há música tocando para pular", delete_after=3)

@bot.tree.command(name="pausar", description="Pausa a música atual")
async def pause(interaction: discord.Interaction):
    voice_client = discord.utils.get(bot.voice_clients, guild=interaction.guild)
    if voice_client and voice_client.is_playing():
        voice_client.pause()
        await interaction.response.send_message("Musica pausada", delete_after=20)
    else:
        await interaction.response.send_message("Não há música tocando para pausar", delete_after=3)

@bot.tree.command(name="continuar", description="Retorna a tocar a música atual")
async def resume(interaction: discord.Interaction):
    voice_client = discord.utils.get(bot.voice_clients, guild=interaction.guild)
    if voice_client and voice_client.is_paused():
        voice_client.resume()
        await interaction.response.send_message("Continuando a tocar", delete_after=3)
    elif voice_client:
        await interaction.response.send_message("Música já está tocando", delete_after=3)
    else:
        await interaction.response.send_message("Não há música tocando", delete_after=3)